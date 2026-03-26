"""
Lector del Tablero_Resumido.xlsx
Estructura verificada contra el archivo real:

TABLERO DIARIO
  Fila 2  → fechas
  Fila 3  → día semana
  Fila 4  → tipo día (LV / SDF)
  Fila 6  → TOTAL KG - TOTAL
  Fila 7  → TOTAL KG - TM
  Fila 8  → TOTAL KG - TT
  Fila 9  → EFICIENCIA % TOTAL
  Fila 10 → EFICIENCIA % TM
  Fila 11 → EFICIENCIA % TT
  Fila 14 → MERMA %
  Fila 15 → DECOMISO %
  Fila 19 → SEMI CPC
  Fila 20 → BALANCÍN CPC
  Fila 21 → CHASIS CPC
  Fila 22 → CAMIONETAS CPC
  Fila 24 → SEMI Ituzaingó
  Fila 25 → BALANCÍN Ituzaingó
  Fila 26 → CHASIS Ituzaingó
  Fila 27 → CAMIONETAS Ituzaingó
  Fila 29 → PUNTAJE TRANSPORTES
  Fila 36 → HORA FIN DESPACHO CPC
  Fila 54 → PRESENTISMO TM
  Fila 55 → PRESENTISMO TT
"""

import os
from datetime import datetime, date
import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "operaciones.xlsx")

_TD_ROW_MAP = {
    3: "dia_semana", 4: "tipo_dia",
    6: "prod_total_kg", 7: "prod_tm_kg", 8: "prod_tt_kg",
    9: "efic_total", 10: "efic_tm", 11: "efic_tt",
    14: "merma", 15: "decomiso",
    19: "semi_cpc", 20: "balancin_cpc", 21: "chasis_cpc", 22: "camioneta_cpc",
    24: "semi_itu", 25: "balancin_itu", 26: "chasis_itu", 27: "camioneta_itu",
    29: "puntaje_transporte",
    36: "hora_fin_despacho",
    54: "presentismo_tm", 55: "presentismo_tt",
}

def _clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s in ("-", "", " ", "\xa0", "#REF!", "#N/A", "#VALUE!") or s.startswith("="):
            return None
    return v

def _pct_efic(v):
    """Eficiencia ya viene como entero (93 = 93%). Si <1, multiplica x100."""
    if v is None: return None
    try:
        f = float(v)
        return round(f, 1) if f > 1 else round(f * 100, 1)
    except: return None

def _pct_merma(v):
    """Merma/Decomiso viene como decimal (0.034 = 3.4%). Multiplica x100."""
    if v is None: return None
    try:
        f = float(v)
        return round(f * 100, 2) if f <= 1 else round(f, 2)
    except: return None

def _pct(v):
    """Genérico: asume decimal si <= 1, entero si > 1."""
    if v is None: return None
    try:
        f = float(v)
        return round(f, 1) if f > 1 else round(f * 100, 1)
    except: return None

def _fmt_date(d):
    if isinstance(d, (datetime, date)):
        return d.strftime("%d/%m/%Y")
    return str(d)

def _open():
    return openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

_cache_dias = None
_cache_meses = None

def invalidar_cache():
    global _cache_dias, _cache_meses
    _cache_dias = None
    _cache_meses = None

def _leer_tablero_diario():
    wb = _open()
    ws = wb["TABLERO DIARIO"]
    rows = {}
    target = set(_TD_ROW_MAP.keys()) | {2}
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=61, values_only=True), start=1):
        if r_idx in target:
            rows[r_idx] = list(row)
    wb.close()

    dias = []
    for col_idx, raw in enumerate(rows.get(2, [])):
        if not isinstance(raw, datetime):
            continue
        def g(r):
            return _clean(rows.get(r, [])[col_idx] if col_idx < len(rows.get(r, [])) else None)
        dia = {"fecha": raw.date(), "fecha_str": _fmt_date(raw)}
        for rn, campo in _TD_ROW_MAP.items():
            dia[campo] = g(rn)
        dia["efic_total_pct"] = _pct_efic(dia["efic_total"])
        dia["efic_tm_pct"]    = _pct_efic(dia["efic_tm"])
        dia["efic_tt_pct"]    = _pct_efic(dia["efic_tt"])
        dia["merma_pct"]      = _pct_merma(dia["merma"])
        dia["decomiso_pct"]   = _pct_merma(dia["decomiso"])
        nums = [dia.get(k) for k in ("semi_cpc","balancin_cpc","chasis_cpc","camioneta_cpc",
                                      "semi_itu","balancin_itu","chasis_itu","camioneta_itu")]
        ok = [n for n in nums if isinstance(n, (int, float))]
        dia["total_camiones"] = int(sum(ok)) if ok else None
        dias.append(dia)
    return dias

def _get_dias():
    global _cache_dias
    if _cache_dias is None:
        _cache_dias = _leer_tablero_diario()
    return _cache_dias

def get_dias_tablero(fecha_inicio=None, fecha_fin=None):
    dias = _get_dias()
    if fecha_inicio:
        dias = [d for d in dias if d["fecha"] >= fecha_inicio]
    if fecha_fin:
        dias = [d for d in dias if d["fecha"] <= fecha_fin]
    return dias

def get_dia(fecha: date):
    r = get_dias_tablero(fecha, fecha)
    return r[0] if r else None

def get_ultimo_dia():
    dias = _get_dias()
    ok = [d for d in dias if isinstance(d.get("prod_total_kg"), (int, float)) and d["prod_total_kg"] > 0]
    return ok[-1] if ok else None

def get_ultimos_n_dias(n=7):
    dias = _get_dias()
    ok = [d for d in dias if isinstance(d.get("prod_total_kg"), (int, float)) and d["prod_total_kg"] > 0]
    return ok[-n:]

def get_resumen_mensual(n_ultimos=None):
    global _cache_meses
    if _cache_meses is None:
        wb = _open()
        ws = wb["MES CERRADO"]
        rows = {}
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
            rows[r_idx] = list(row)
        wb.close()
        meses = []
        for col_idx, v in enumerate(rows.get(2, [])):
            if not isinstance(v, datetime):
                continue
            def g(r):
                rdata = rows.get(r, [])
                val = rdata[col_idx] if col_idx < len(rdata) else None
                return _clean(val)
            pr = g(3)
            if pr is None:
                continue
            meses.append({
                "mes": v.strftime("%b %Y"),
                "mes_largo": v.strftime("%B %Y"),
                "fecha": v.date(),
                "prod_real_tn": round(float(pr), 1),
                "prod_obj_tn":  round(float(g(4)), 1) if g(4) else None,
                "pct_vs_obj":   (round(float(g(5))*100, 1) if g(5) is not None else None),
                "efic_pct":     _pct_efic(g(9)),
                "merma_pct":    _pct_merma(g(10)),
                "costo_kg":     g(27),
            })
        _cache_meses = meses
    return _cache_meses[-n_ultimos:] if n_ultimos else _cache_meses

def get_temas(sector=None, solo_abiertos=True):
    wb = _open()
    ws = wb["TEMAS"]
    temas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        tema_txt = row[4]
        if not tema_txt or (isinstance(tema_txt, str) and tema_txt.startswith("=")):
            continue
        cerrado = str(row[12]).upper() == "SI" if row[12] else False
        item = {
            "sector": row[1], "responsable": row[2], "prioridad": row[3],
            "tema": tema_txt, "avance": row[11], "cerrado": cerrado,
        }
        if sector and row[1] and sector.upper() not in str(row[1]).upper():
            continue
        if solo_abiertos and cerrado:
            continue
        temas.append(item)
    wb.close()
    return temas

def get_flota():
    wb = _open()
    ws = wb["LOGÍSTICA"]
    flota = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        flota.append({"patente": row[0], "viajes_sem": row[1],
                      "transporte": row[4], "capacidad": row[5],
                      "gps": row[6], "pala": row[8], "modelo": row[9]})
    wb.close()
    return flota

def get_devoluciones():
    wb = _open()
    ws = wb["DEVOLUCIONES - anterior"]
    devs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]: continue
        devs.append({"fecha": _fmt_date(row[1]), "codigo": row[2],
                     "descripcion": row[3], "cliente": row[5],
                     "cajas": row[6], "motivo": row[7]})
    wb.close()
    return devs

def get_rechazos():
    wb = _open()
    ws = wb["RECHAZOS - anterior"]
    rechazos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]: continue
        rechazos.append({"fecha": _fmt_date(row[1]), "mes": row[2],
                         "codigo": row[3], "descripcion": row[4],
                         "cliente": row[5], "cajas": row[6],
                         "responsable": row[7], "motivo": row[8]})
    wb.close()
    return rechazos

def get_recortes():
    wb = _open()
    ws = wb["RECORTES"]
    recortes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]: continue
        recortes.append({"fecha": _fmt_date(row[1]), "codigo": row[3],
                         "descripcion": row[4], "cajas": row[6],
                         "kg": row[7], "motivo": row[8], "otro_dia": row[10]})
    wb.close()
    return recortes

def get_contexto_completo(pregunta_usuario: str = "") -> str:
    p = pregunta_usuario.lower()
    partes = []

    # Siempre: último día
    u = get_ultimo_dia()
    if u:
        s=u["semi_cpc"] or 0; b=u["balancin_cpc"] or 0; c=u["chasis_cpc"] or 0; cam=u["camioneta_cpc"] or 0
        si=u["semi_itu"] or 0; bi=u["balancin_itu"] or 0; ci=u["chasis_itu"] or 0; cai=u["camioneta_itu"] or 0
        partes.append(
f"""=== ÚLTIMO DÍA CON PRODUCCIÓN: {u['fecha_str']} ({u.get('dia_semana','-')}) - Tipo: {u.get('tipo_dia','-')} ===
Producción total : {u['prod_total_kg']:>10,.0f} kg
  Turno Mañana   : {(u['prod_tm_kg'] or 0):>10,.0f} kg
  Turno Tarde    : {(u['prod_tt_kg'] or 0):>10,.0f} kg
Eficiencia       : Total={u['efic_total_pct']}%  TM={u['efic_tm_pct']}%  TT={u['efic_tt_pct']}%
Merma            : {u['merma_pct']}%   Decomiso: {u['decomiso_pct']}%
Presentismo      : TM={u.get('presentismo_tm','-')} personas  TT={u.get('presentismo_tt','-')} personas
Camiones CPC+Dep : Semi={s}  Balancín={b}  Chasis={c}  Camioneta={cam}  → Subtotal={s+b+c+cam}
Camiones Itu     : Semi={si}  Balancín={bi}  Chasis={ci}  Camioneta={cai}  → Subtotal={si+bi+ci+cai}
Total camiones   : {u['total_camiones']}
Puntaje transporte: {u.get('puntaje_transporte', '-')}
Hora fin despacho : {u.get('hora_fin_despacho', '-')}""")

    # Últimos 7 días
    if any(w in p for w in ["semana","últimos","tendencia","promedio","días","dias","varios"]):
        dias = get_ultimos_n_dias(7)
        if dias:
            lineas = [
                f"  {d['fecha_str']} {d.get('dia_semana',''):3} | "
                f"Prod: {d['prod_total_kg']:>7,.0f} kg | "
                f"Efic: {d['efic_total_pct']}% (TM:{d['efic_tm_pct']}% TT:{d['efic_tt_pct']}%) | "
                f"Merma: {d['merma_pct']}% | "
                f"Pres: TM={d.get('presentismo_tm','-')} TT={d.get('presentismo_tt','-')} | "
                f"Camiones: {d['total_camiones']}"
                for d in dias
            ]
            partes.append("=== ÚLTIMOS 7 DÍAS CON PRODUCCIÓN ===\n" + "\n".join(lineas))

    # Datos mensuales
    kw_mes = ["mes","mensual","enero","febrero","marzo","abril","mayo","junio","julio","agosto",
              "septiembre","octubre","noviembre","diciembre","histórico","historico",
              "año","2025","2024","2026","objetivo","plan","costo"]
    if any(w in p for w in kw_mes):
        meses = get_resumen_mensual(n_ultimos=12)
        lineas = []
        for m in meses:
            cumpl = f"{m['pct_vs_obj']}%" if m['pct_vs_obj'] else "-"
            costo = f"${m['costo_kg']:,.0f}" if m['costo_kg'] else "-"
            lineas.append(
                f"  {m['mes']:>8} | Prod: {m['prod_real_tn']:>8} TN | "
                f"Obj: {m['prod_obj_tn']:>8} TN | Cumpl: {cumpl:>6} | "
                f"Efic: {m['efic_pct']}% | Merma: {m['merma_pct']}% | Costo/kg: {costo}"
            )
        partes.append("=== RESUMEN MENSUAL (últimos 12 meses) ===\n" + "\n".join(lineas))

    # Presentismo
    if any(w in p for w in ["presentismo","personas","operarios","dotación","dotacion","gente","línea","linea"]):
        dias = get_ultimos_n_dias(7)
        lineas = [
            f"  {d['fecha_str']} {d.get('dia_semana',''):3} | TM: {d.get('presentismo_tm','-')} personas | TT: {d.get('presentismo_tt','-')} personas"
            for d in dias
        ]
        partes.append("=== PRESENTISMO ÚLTIMOS 7 DÍAS ===\n" + "\n".join(lineas))

    # Camiones
    if any(w in p for w in ["camion","camión","semi","balancín","balancin","chasis","camioneta",
                             "flota","logística","logistica","transporte","viaje","despacho"]):
        flota = get_flota()
        caps: dict = {}
        for v in flota:
            cap = str(v.get("capacidad") or "desconocida")
            caps[cap] = caps.get(cap, 0) + 1
        partes.append(
            f"=== FLOTA TOTAL: {len(flota)} vehículos activos ===\n" +
            "\n".join(f"  Capacidad {k}: {n} vehículo(s)" for k, n in sorted(caps.items()))
        )

    # Temas
    if any(w in p for w in ["tema","pendiente","acción","accion","problema","abierto",
                             "calidad","higiene","abastecimiento","produccion","producción"]):
        temas = get_temas(solo_abiertos=True)
        if temas:
            lineas = [
                f"  [{t['sector']}] {t['tema']} | Resp: {t['responsable']} | "
                f"Avance: {int((t['avance'] or 0)*100)}%"
                for t in temas[:12]
            ]
            partes.append(f"=== TEMAS ABIERTOS ({len(temas)} total) ===\n" + "\n".join(lineas))

    # Devoluciones
    if any(w in p for w in ["devolución","devolucion","rechazo","rechaz","cliente","no recibido"]):
        devs = get_devoluciones()[-6:]
        if devs:
            lineas = [f"  {d['fecha']} | {str(d['cliente'])[:40]} | {str(d['descripcion'])[:30]} | {d['cajas']} cajas | {d['motivo']}" for d in devs]
            partes.append("=== ÚLTIMAS DEVOLUCIONES ===\n" + "\n".join(lineas))
        rechazos = get_rechazos()[-6:]
        if rechazos:
            lineas = [f"  {r['fecha']} | {str(r['cliente'])[:40]} | {str(r['descripcion'])[:30]} | {r['cajas']} cajas | {r['motivo']}" for r in rechazos]
            partes.append("=== ÚLTIMOS RECHAZOS ===\n" + "\n".join(lineas))

    # Recortes
    if any(w in p for w in ["recorte","no se llegó","no se llego","faltante","discontinu","no se hizo"]):
        recortes = get_recortes()[-6:]
        if recortes:
            lineas = [f"  {r['fecha']} | {str(r['descripcion'] or '')[:35]} | {r['kg']} kg | {r['motivo']}" for r in recortes]
            partes.append("=== ÚLTIMOS RECORTES DE PRODUCCIÓN ===\n" + "\n".join(lineas))

    return "\n\n".join(partes) if partes else "No hay datos cargados en el Excel todavía."
